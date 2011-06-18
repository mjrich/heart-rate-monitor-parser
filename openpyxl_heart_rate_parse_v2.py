#!/usr/bin/python
import os, sys #standard python libraries for general system and operating system specific
import simplejson
import mx.DateTime
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook

# Need to turn into function and then increment accross as many files as needed

def scrapesheet(link):
	wb2 = load_workbook(filename = link) # Make sure file name does not start with number!!! 
	ws2 = wb2.get_active_sheet() # Make sure sheet name does not start with number!!! 
	datetime_text = ws2.cell('B3').value # This is where the time stamp starts
	dt = mx.DateTime.DateTimeFrom(datetime_text)
	#print dt
	dataset = []
	for row in ws2.range('U3:U50000'):
		for cell in row:
			if cell.value == None:
				continue
			#print dt, cell.value
			dataset += [(str(dt), cell.value)]
			dt = dt + mx.DateTime.DateTimeDeltaFromSeconds(2)
	return dataset

datadump = []
nums = ['4', '5']
for num in nums:
	location = 'C:\XXX\%s.xlsx' % num # Directories and files should be specified here before running.
	datadump += scrapesheet(location)

simplejson.dump(datadump, open(sys.argv[1],'w')) # This is the expression for exporting the data.  Name it when running the script.